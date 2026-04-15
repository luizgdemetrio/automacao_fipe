""" """

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# CONFIGURAÇÕES — ajuste aqui se algo mudar
# ─────────────────────────────────────────────
COLUNA_PLACA = "Placa"
COLUNA_FIPE = "FIPE"
COLUNA_FIPE_REF = "FIPE REF"


# ─────────────────────────────────────────────
# FUNÇÃO: carregar workbook e aba ativa
# ─────────────────────────────────────────────
def carregar_planilha(caminho_arquivo: str):
    """
    Abre o arquivo .xlsx e retorna o workbook e a aba ativa.

    Parâmetros:
        caminho_arquivo: caminho completo do arquivo, ex: r"C:\Downloads\ffr.xlsx"

    Retorna:
        wb  — o workbook (arquivo inteiro)
        ws  — a aba ativa (worksheet)
    """
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active
    return wb, ws


# ─────────────────────────────────────────────
# FUNÇÃO: mapear cabeçalhos → número da coluna
# ─────────────────────────────────────────────
def mapear_cabecalhos(ws) -> dict:
    """
    Lê a linha 1 e devolve um dicionário:
        { "Nome do cabeçalho": número_da_coluna, ... }

    Exemplo de retorno:
        { "Placa": 3, "FIPE": 15, "FIPE REF": 16 }
    """
    cabecalhos = {}
    for celula in ws[1]:
        if celula.value is not None:
            cabecalhos[str(celula.value).strip()] = celula.column
    return cabecalhos


# ─────────────────────────────────────────────
# FUNÇÃO: garantir que as colunas FIPE existem
# ─────────────────────────────────────────────
def garantir_colunas_fipe(ws, cabecalhos: dict) -> dict:
    """
    Verifica se as colunas "FIPE" e "FIPE REF" já existem.
    Se não existirem, cria no final da planilha com cabeçalho estilizado.

    Retorna o dicionário de cabeçalhos atualizado.
    """
    ultima_coluna = ws.max_column

    # Estilo do cabeçalho das novas colunas
    fonte_negrito = Font(bold=True)
    fundo_amarelo = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    centralizado = Alignment(horizontal="center")

    for nome_coluna in [COLUNA_FIPE, COLUNA_FIPE_REF]:
        if nome_coluna not in cabecalhos:
            ultima_coluna += 1
            celula = ws.cell(row=1, column=ultima_coluna, value=nome_coluna)
            celula.font = fonte_negrito
            celula.fill = fundo_amarelo
            celula.alignment = centralizado
            cabecalhos[nome_coluna] = ultima_coluna
            print(
                f"  [+] Coluna '{nome_coluna}' criada na coluna {get_column_letter(ultima_coluna)}"
            )
        else:
            print(f"  [=] Coluna '{nome_coluna}' já existe — mantida.")

    return cabecalhos


# ─────────────────────────────────────────────
# FUNÇÃO: listar linhas sem valor na FIPE
# ─────────────────────────────────────────────
def listar_linhas_sem_fipe(ws, cabecalhos: dict) -> list:
    """
    Percorre todas as linhas de dados (a partir da linha 2)
    e retorna uma lista de dicionários com as linhas que
    ainda não possuem valor na coluna FIPE.

    Cada item da lista contém:
        {
            "linha":  número da linha no Excel (int),
            "placa":  valor da coluna Placa (str),
        }
    """
    col_placa = cabecalhos.get(COLUNA_PLACA)
    col_fipe = cabecalhos.get(COLUNA_FIPE)

    if col_placa is None:
        raise ValueError(f"Coluna '{COLUNA_PLACA}' não encontrada na planilha.")
    if col_fipe is None:
        raise ValueError(
            f"Coluna '{COLUNA_FIPE}' não encontrada. Execute garantir_colunas_fipe() primeiro."
        )

    pendentes = []

    for linha in ws.iter_rows(min_row=2, values_only=False):
        numero_linha = linha[0].row
        placa = linha[col_placa - 1].value
        valor_fipe = linha[col_fipe - 1].value

        # Ignora linhas sem placa
        if not placa:
            continue

        # Coleta se FIPE estiver vazia
        if valor_fipe is None or str(valor_fipe).strip() == "":
            pendentes.append(
                {
                    "linha": numero_linha,
                    "placa": str(placa).strip().upper(),
                }
            )

    return pendentes


# ─────────────────────────────────────────────
# FUNÇÃO: gravar FIPE e FIPE REF em uma linha
# ─────────────────────────────────────────────
def gravar_fipe(
    ws, cabecalhos: dict, numero_linha: int, valor_fipe: float, mes_referencia: str
):
    """
    Grava o valor da FIPE e o mês de referência na linha indicada.

    Parâmetros:
        ws              — worksheet aberta
        cabecalhos      — dicionário de cabeçalhos mapeados
        numero_linha    — número da linha no Excel onde gravar
        valor_fipe      — valor numérico, ex: 45780.00
        mes_referencia  — texto do mês, ex: "março de 2025"
    """
    col_fipe = cabecalhos[COLUNA_FIPE]
    col_fipe_ref = cabecalhos[COLUNA_FIPE_REF]

    # Grava o valor como número com formato de moeda
    celula_valor = ws.cell(row=numero_linha, column=col_fipe, value=valor_fipe)
    celula_valor.number_format = "R$ #,##0.00"

    # Grava o mês de referência como texto
    ws.cell(row=numero_linha, column=col_fipe_ref, value=mes_referencia)


# ─────────────────────────────────────────────
# FUNÇÃO: salvar o arquivo
# ─────────────────────────────────────────────
def salvar_planilha(wb, caminho_arquivo: str):
    """
    Salva o workbook no mesmo caminho do arquivo original.
    """
    wb.save(caminho_arquivo)
    print(f"  [✓] Planilha salva em: {caminho_arquivo}")


# ─────────────────────────────────────────────
# EXECUÇÃO DIRETA — teste rápido do módulo
# ─────────────────────────────────────────────
if __name__ == "__main__":

    # Substitua pelo caminho real do seu arquivo para testar
    ARQUIVO = r"FFR_Novo.xlsx"

    print("=" * 50)
    print("Testando planilha.py")
    print("=" * 50)

    # 1. Carrega
    wb, ws = carregar_planilha(ARQUIVO)
    print(f"\n[1] Planilha carregada — {ws.max_row - 1} linha(s) de dados\n")

    # 2. Mapeia cabeçalhos
    cabecalhos = mapear_cabecalhos(ws)
    print(f"[2] Cabeçalhos encontrados:\n    {list(cabecalhos.keys())}\n")

    # 3. Garante colunas FIPE e FIPE REF
    print("[3] Verificando colunas FIPE:")
    cabecalhos = garantir_colunas_fipe(ws, cabecalhos)

    # 4. Lista pendentes
    pendentes = listar_linhas_sem_fipe(ws, cabecalhos)
    print(f"\n[4] Linhas sem FIPE: {len(pendentes)}")
    for p in pendentes[:5]:  # mostra as 5 primeiras
        print(f"    Linha {p['linha']} — Placa: {p['placa']}")
    if len(pendentes) > 5:
        print(f"    ... e mais {len(pendentes) - 5} linha(s)")

    # 5. Salva (já com as colunas criadas se não existiam)
    salvar_planilha(wb, ARQUIVO)
